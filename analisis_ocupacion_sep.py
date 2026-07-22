"""
Análisis de Ocupación Logística — un solo archivo de stock físico.

Entradas:
  stock.xlsx  → stock por SKU (Bodega, Zona, Formato, inv Fisico, Capacidad Utilizada)
  capacidades.xlsx (opcional) → template de capacidades por bodega/zona

Salida: reporte_ocupacion_sep.html + .xlsx
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import re
import unicodedata
import argparse
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import warnings
from pathlib import Path

# ── Flexible column mapping ───────────────────────────────────────────────────
def _norm(s):
    """Normalize: strip accents, lowercase, remove separators."""
    s = unicodedata.normalize("NFKD", str(s).strip())
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[\s_\-\./]+", "", s).lower()

_COL_ALIASES = {
    "Codigo":              ["codigo","code","sku","cod","codarticulo","codigoproducto","idproducto"],
    "Articulo":            ["articulo","descripcion","description","desc","nombre","producto","nombrearti"],
    "Bodega":              ["bodega","almacen","camara","warehouse","deposito","storage"],
    "Zona":                ["zona","nivel","area","zone","level","piso"],
    "ID_Posición":         ["idposicion","idposicion","posicion","idpos","posicionid","position","ubicacion","id_posicion"],
    "Formato":             ["formato","format","tipocaja","tipoembalaje","presentacion","embalaje","caja"],
    "Cantidad":            ["cantidad","qty","quantity","cant","unidades","units"],
    "inv Fisico":          ["invfisico","inventariofisico","stockfisico","invfis","fisico","stockreal","inventario"],
    "Capacidad Utilizada": ["capacidadutilizada","caputilizada","caputil","utilizacion","ocupacion","capusada"],
}
_ALIAS_LOOKUP = {alias: std for std, aliases in _COL_ALIASES.items() for alias in aliases}

def map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename df columns to standard names using fuzzy matching."""
    rename = {}
    for col in df.columns:
        key = _norm(col)
        if key in _ALIAS_LOOKUP and col != _ALIAS_LOOKUP[key]:
            rename[col] = _ALIAS_LOOKUP[key]
    return df.rename(columns=rename) if rename else df

warnings.filterwarnings("ignore")

# ─── CONFIG (defaults, sobreescritos por CLI args) ────────────────────────────
INPUT_FILE       = "maestra de codigosv2.xlsx"
CAPACIDADES_FILE = None   # None = leer desde inf del INPUT_FILE
OUTPUT_XLSX      = "reporte_ocupacion_sep.xlsx"
OUTPUT_HTML      = "reporte_ocupacion_sep.html"
LABEL            = ""     # etiqueta libre (ej: "Promedio Semanal Jul-Sep")
OUTPUT_DIR       = None   # si se especifica, los outputs van a ese directorio

TH_GREEN  = 0.75
TH_YELLOW = 0.85

# Mapeo Cámara (inf) → Bodega (stock file) — fallback cuando no hay template
CAMARA_A_BODEGA = {
    "Camara 5": "Congelado salado 5",
    "Camara 4": "Mantencion salado 4",
}

FILL_GREEN  = PatternFill("solid", fgColor="00B050")
FILL_YELLOW = PatternFill("solid", fgColor="FFBF00")
FILL_RED    = PatternFill("solid", fgColor="FF0000")
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER = Font(bold=True, color="FFFFFF")


# ─── HELPERS ──────────────────────────────────────────────────────────────────
def semaforo_fill(pct):
    if pct is None or (isinstance(pct, float) and np.isnan(pct)):
        return None
    if pct < TH_GREEN:
        return FILL_GREEN
    elif pct <= TH_YELLOW:
        return FILL_YELLOW
    return FILL_RED


def semaforo_label(pct) -> str:
    if pct is None or (isinstance(pct, float) and np.isnan(pct)):
        return "Sin Cap."
    if pct < TH_GREEN:
        return "OK"
    elif pct <= TH_YELLOW:
        return "ALERTA"
    return "CRITICO"


def semaforo_css(pct) -> str:
    if pct is None or (isinstance(pct, float) and np.isnan(pct)):
        return "cell-nocap"
    if pct < TH_GREEN:
        return "cell-green"
    elif pct <= TH_YELLOW:
        return "cell-yellow"
    return "cell-red"


# ─── PASO 1: STOCK FÍSICO (BASE) ─────────────────────────────────────────────
def load_sku_master(xls: pd.ExcelFile):
    """
    Lee pestaña BASE del archivo de stock físico.
    Columnas requeridas: Codigo, Bodega, Zona, Formato, Capacidad Utilizada
    pos_vol = Capacidad Utilizada (posiciones físicas ocupadas por SKU).
    Devuelve (stock_df, factor_map) donde factor_map es por compatibilidad.
    """
    # header=0 (v2); fallback header=1 (v1)
    df = xls.parse("BASE", header=0)
    df.columns = df.columns.str.strip()
    df = map_columns(df)
    if "Codigo" not in df.columns:
        df = xls.parse("BASE", header=1)
        df.columns = df.columns.str.strip()
        df = map_columns(df)

    df["Codigo"] = df["Codigo"].astype(str).str.strip()
    df = df[df["Codigo"].str.match(r"^\d+$")].copy()

    for col in ["Capacidad Utilizada", "inv Fisico", "Cantidad"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["pos_vol"] = df.get("Capacidad Utilizada", pd.Series(0.0, index=df.index))

    DETAIL_COLS = ["Codigo", "Articulo", "Bodega", "Zona", "ID_Posición",
                   "Formato", "Cantidad", "inv Fisico", "Capacidad Utilizada", "pos_vol"]
    keep = [c for c in DETAIL_COLS if c in df.columns]
    stock = df[keep].copy()

    sin_bodega = stock["Bodega"].isna().sum()
    print(f"      {len(stock)} SKUs | {sin_bodega} sin bodega")

    return stock, {}


# ─── PASO 3: CAPACIDAD REAL (inf o template_capacidades) ────────────────────
def load_capacity(xls: pd.ExcelFile) -> pd.DataFrame:
    """
    Retorna DataFrame con columnas [Bodega, Zona, Cap_Real].
    Zona ∈ {"Piso", "Niveles 1–3", "Nivel 4"}.

    Fuente preferida: template_capacidades.xlsx (nuevo formato simplificado).
    Fallback: pestaña inf del maestro (posiciones fijas).

    Nuevo template:
      Bodega | Zona | ID_Posición | Num_Slots | Cap. Nominal Bandejas | Cap. Nominal Racks
      Zona física = Rack → capacidad en racks, Nivel derivado de ID_Posición
      Zona física = Piso → capacidad en bandejas, Nivel = Piso
    """
    def _idpos_to_zona(id_pos: str) -> str:
        """ID que empieza con P → 'Piso'; cualquier N (rack) → 'Rack'."""
        s = str(id_pos).upper().strip()
        return "Piso" if re.match(r"^P", s) else "Rack"

    # ── Fuente externa: template_capacidades.xlsx (nuevo formato) ─────────────
    if CAPACIDADES_FILE and Path(CAPACIDADES_FILE).exists():
        print(f"  Capacidades desde: {CAPACIDADES_FILE}")
        df = pd.read_excel(CAPACIDADES_FILE, sheet_name="Capacidades", header=2)
        df.columns = df.columns.str.strip()

        # Mapear columnas por palabra clave (tolerante a saltos de línea en header)
        rename = {}
        for col in df.columns:
            cl = col.lower().replace("\n", " ")
            if "bodega" in cl:                            rename[col] = "Bodega"
            elif "zona" in cl:                            rename[col] = "Zona_Fis"
            elif "posici" in cl or "id_pos" in cl:       rename[col] = "ID_Pos"
            elif "bandeja" in cl or "band" in cl:        rename[col] = "Cap_Band"
            elif "rack" in cl:                            rename[col] = "Cap_Rack"
        df = df.rename(columns=rename)

        # Eliminar fila de totales y filas sin bodega
        df = df[df.get("Bodega", pd.Series()).notna()]
        df = df[~df["Bodega"].astype(str).str.upper().str.contains("TOTAL|—")]

        df["Bodega"] = df["Bodega"].astype(str).str.strip()
        df["Bodega"] = df["Bodega"].astype(str).str.strip()

        df["Zona"] = df["ID_Pos"].apply(_idpos_to_zona)

        df["Cap_Band"] = pd.to_numeric(df.get("Cap_Band", pd.Series(dtype=float)),
                                        errors="coerce").fillna(0)
        df["Cap_Rack"] = pd.to_numeric(df.get("Cap_Rack", pd.Series(dtype=float)),
                                        errors="coerce").fillna(0)

        def _cap(row):
            return row["Cap_Band"] if row["Zona"] == "Piso" else row["Cap_Rack"]

        df["Cap_Real"] = df.apply(_cap, axis=1)
        df = df[df["Cap_Real"] > 0]

        return df.groupby(["Bodega", "Zona"])["Cap_Real"].sum().reset_index()

    # ── Fallback: pestaña inf (posiciones hardcodeadas) ───────────────────────
    raw = xls.parse("inf", header=None)

    def val(row, col):
        try:
            v = raw.iloc[row, col]
            return float(v) if pd.notna(v) else 0.0
        except Exception:
            return 0.0

    # Piso = suma Bandeja S + M + L; Racks totales asignados al nivel con más SKUs
    capacidades = [
        (CAMARA_A_BODEGA["Camara 5"], "Niveles 1–3", val(9, 2)),
        (CAMARA_A_BODEGA["Camara 5"], "Nivel 4",     val(9, 2)),
        (CAMARA_A_BODEGA["Camara 5"], "Piso",        val(23, 2) + val(24, 2) + val(25, 2)),
        (CAMARA_A_BODEGA["Camara 4"], "Niveles 1–3", val(9, 5)),
        (CAMARA_A_BODEGA["Camara 4"], "Nivel 4",     val(9, 5)),
        (CAMARA_A_BODEGA["Camara 4"], "Piso",        val(23, 5) + val(24, 5) + val(25, 5)),
    ]
    capacity = pd.DataFrame(capacidades, columns=["Bodega", "Zona", "Cap_Real"])
    return capacity.groupby(["Bodega", "Zona"])["Cap_Real"].max().reset_index()


# ─── PASO 3b: CONSOLIDAR ─────────────────────────────────────────────────────
def consolidate(plan: pd.DataFrame, capacity: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa por [Bodega, Zona] donde Zona = "Piso" o "Rack" (de BASE v2)
    o derivado de Nivel en v1 ("Piso"/"Nivel 4"/"Niveles 1–3").
    Toda demanda Piso (Bandeja S/M/L) comparte capacidad piso.
    Toda demanda Rack (cualquier formato en rack) comparte capacidad rack.
    """
    p = plan.dropna(subset=["Bodega"]).copy()
    if "Zona" not in p.columns:
        def _zona(nivel):
            if str(nivel) == "Piso": return "Piso"
            return "Rack"
        p["Zona"] = p.get("Nivel", pd.Series("Piso", index=p.index)).apply(_zona)

    grp_cols = ["Bodega", "Zona"]
    if "Formato" in p.columns:
        grp_cols.append("Formato")

    grouped = (
        p.groupby(grp_cols, dropna=False)
        .agg(Pos_Vol=("pos_vol", "sum"))
        .reset_index()
    )

    result = grouped.merge(capacity, on=["Bodega", "Zona"], how="left")
    result["Cap_Real"] = result["Cap_Real"].fillna(0)

    def pct(num, den):
        return num / den if den > 0 else np.nan

    result["Pct_Vol"] = result.apply(lambda r: pct(r["Pos_Vol"], r["Cap_Real"]), axis=1)

    sort_cols = ["Bodega", "Zona"] + (["Formato"] if "Formato" in result.columns else [])
    return result.sort_values(sort_cols).reset_index(drop=True)


# ─── PASO 4a: OUTPUT EXCEL ───────────────────────────────────────────────────
def write_excel(result: pd.DataFrame, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Ocupacion SEP"

    has_fmt_col = "Formato" in result.columns
    headers = [
        "Bodega / Cámara", "Zona",
        *( ["Formato"] if has_fmt_col else [] ),
        f"Pos. Requeridas\n({LABEL or 'Stock'})",
        "Capacidad Real\n(Posiciones)",
        f"% Ocupación\n({LABEL or 'Stock'})",
        "Estado",
    ]

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill      = FILL_HEADER
        c.font      = FONT_HEADER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 45

    pct_col    = 5 if has_fmt_col else 4
    estado_col = pct_col + 1

    for ri, row in enumerate(result.itertuples(index=False), 2):
        pv   = row.Pct_Vol
        fmt_val = getattr(row, "Formato", None)
        vals = [
            row.Bodega, row.Zona,
            *( [fmt_val] if has_fmt_col else [] ),
            round(row.Pos_Vol, 1), round(row.Cap_Real, 1),
            pv, semaforo_label(pv),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if ci == pct_col:
                c.number_format = "0.0%"
                f = semaforo_fill(pv)
                if f: c.fill = f
            elif ci == estado_col:
                f = semaforo_fill(pv)
                if f: c.fill = f
                c.font = Font(bold=True)

    col_widths = [26, 16] + ([18] if has_fmt_col else []) + [22, 22, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Leyenda
    wl = wb.create_sheet("Leyenda")
    for ci, h in enumerate(["Color", "Rango", "Interpretación"], 1):
        c = wl.cell(1, ci, h)
        c.fill = FILL_HEADER; c.font = FONT_HEADER
    rows_ley = [
        (FILL_GREEN,  "Verde",    "< 75%",    "Capacidad disponible"),
        (FILL_YELLOW, "Amarillo", "75%–85%",  "Zona de alerta"),
        (FILL_RED,    "Rojo",     "> 85%",    "Cuello de botella"),
    ]
    for ri, (fill, color, rango, desc) in enumerate(rows_ley, 2):
        wl.cell(ri, 1, color).fill = fill
        wl.cell(ri, 2, rango)
        wl.cell(ri, 3, desc)
    wl.column_dimensions["A"].width = 14
    wl.column_dimensions["B"].width = 12
    wl.column_dimensions["C"].width = 40

    wb.save(path)
    print(f"  Excel → {path}")


# SVG icons (Heroicons outline, 20px viewBox)
_ICON_WAREHOUSE = '<svg xmlns="http://www.w3.org/2000/svg" width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.75" stroke-linecap="round" stroke-linejoin="round"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>'
_ICON_ALERT     = '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M10.29 3.86L1.82 18a2 2 0 0 0 1.71 3h16.94a2 2 0 0 0 1.71-3L13.71 3.86a2 2 0 0 0-3.42 0z"/><line x1="12" y1="9" x2="12" y2="13"/><line x1="12" y1="17" x2="12.01" y2="17"/></svg>'
_ICON_CHECK     = '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>'
_ICON_BOX       = '<svg xmlns="http://www.w3.org/2000/svg" width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.75" stroke-linecap="round" stroke-linejoin="round"><path d="M21 16V8a2 2 0 0 0-1-1.73l-7-4a2 2 0 0 0-2 0l-7 4A2 2 0 0 0 3 8v8a2 2 0 0 0 1 1.73l7 4a2 2 0 0 0 2 0l7-4A2 2 0 0 0 21 16z"/><polyline points="3.27 6.96 12 12.01 20.73 6.96"/><line x1="12" y1="22.08" x2="12" y2="12"/></svg>'
_ICON_INFO      = '<svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><circle cx="12" cy="12" r="10"/><line x1="12" y1="8" x2="12" y2="12"/><line x1="12" y1="16" x2="12.01" y2="16"/></svg>'


# ─── PASO 4b: OUTPUT HTML ────────────────────────────────────────────────────
def write_html(result: pd.DataFrame, plan: pd.DataFrame, path: str):
    total_skus = plan["Codigo"].nunique()
    sin_match  = plan["Bodega"].isna().sum()
    n_critico  = int((result["Pct_Vol"].fillna(0) > TH_YELLOW).sum())
    n_alerta   = int(((result["Pct_Vol"].fillna(0) >= TH_GREEN) &
                      (result["Pct_Vol"].fillna(0) <= TH_YELLOW)).sum())

    def fmt_pct(v):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        return f"{v:.1%}"

    def badge_worst(pm, pp=None):
        vals = [v for v in [pm, pp] if v is not None and not (isinstance(v, float) and np.isnan(v))]
        if not vals:
            return "badge-gray", "Sin Cap."
        worst = max(vals)
        if worst > TH_YELLOW:
            return "badge-red", "Crítico"
        elif worst >= TH_GREEN:
            return "badge-yellow", "Alerta"
        return "badge-green", "OK"

    def gauge_bar(pct, css):
        """Mini barra de progreso inline para la celda de %."""
        if pct is None or (isinstance(pct, float) and np.isnan(pct)):
            return ""
        w = min(pct * 100, 100)
        return f'<div class="bar-wrap"><div class="bar {css}" style="width:{w:.1f}%"></div></div>'

    # pre-index plan by bodega for detail tables
    plan_by_bodega = {}
    if plan is not None and "Bodega" in plan.columns:
        for b, g in plan.groupby("Bodega", dropna=False):
            plan_by_bodega[b] = g

    def _v(row, col, fmt=None):
        if col not in row.index:
            return "—"
        v = row[col]
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "—"
        if fmt == "num":
            return f"{v:,.0f}"
        if fmt == "dec":
            return f"{v:,.2f}"
        return str(v)

    def _zona_chip(zona_str):
        is_rack = str(zona_str) in ("Rack", "Niveles 1–3", "Nivel 4")
        bg  = "#DBEAFE" if is_rack else "#DCFCE7"
        col = "#1D4ED8" if is_rack else "#15803D"
        lbl = "Rack"   if is_rack else "Piso"
        return (
            f'<span class="zona-chip" style="background:{bg};color:{col}">{lbl}</span>'
            f'<span class="zona-label">{zona_str}</span>'
        )

    # ── Construir secciones por Bodega ──
    bodegas_list = [b for b in result["Bodega"].dropna().unique()]
    filter_chips = "".join(
        f'<button class="filter-chip" data-bodega="{b}" onclick="filterCam(this,\'{b}\')">{b}</button>'
        for b in bodegas_list
    )

    sections_html = ""
    for bodega, grp in result.groupby("Bodega", dropna=False):
        bodega_label = bodega if pd.notna(bodega) else "Sin Bodega"
        bodega_id    = str(bodega_label).replace(" ", "_").replace("/", "-")

        grp_critico = int((grp["Pct_Vol"].fillna(0) > TH_YELLOW).sum())
        grp_ok      = int(((grp["Pct_Vol"].fillna(0) < TH_GREEN) &
                           grp["Pct_Vol"].notna()).sum())
        total_cap   = grp["Cap_Real"].sum()

        worst_pv = grp["Pct_Vol"].max() if grp["Pct_Vol"].notna().any() else np.nan
        badge_cls, badge_txt = badge_worst(worst_pv, None)

        crit_icon = f'<span class="stat-icon icon-red">{_ICON_ALERT}</span>' if grp_critico > 0 else ''
        ok_icon   = f'<span class="stat-icon icon-green">{_ICON_CHECK}</span>'

        # ── Summary + inline detail rows per Zona ──
        has_fmt  = "Formato" in grp.columns
        sku_df   = plan_by_bodega.get(bodega, pd.DataFrame())
        n_cols   = 5 + (1 if has_fmt else 0)  # Zona [Formato] Pos Cap Pct Estado
        rows = ""
        row_idx  = 0
        for _, row in grp.iterrows():
            pv       = row["Pct_Vol"]
            css      = semaforo_css(pv)
            zona_str = str(row["Zona"])
            fmt_val  = _v(row, "Formato") if has_fmt else ""
            rid      = f"{bodega_id}_{row_idx}"
            row_idx += 1

            fmt_cell = f'<td class="td-left td-mono">{fmt_val}</td>' if has_fmt else ""

            # Filter SKUs for this zona (and formato if grouped by it)
            if not sku_df.empty and "Zona" in sku_df.columns:
                mask = sku_df["Zona"] == zona_str
                if has_fmt and "Formato" in sku_df.columns and fmt_val not in ("—", ""):
                    mask = mask & (sku_df["Formato"] == fmt_val)
                zone_skus = sku_df[mask]
            else:
                zone_skus = pd.DataFrame()

            sku_count = len(zone_skus)
            sku_rows  = ""
            for _, sk in zone_skus.iterrows():
                sku_rows += f"""
                <tr class="dt-row">
                  <td class="dt-code">{_v(sk,'Codigo')}</td>
                  <td class="dt-desc">{_v(sk,'Articulo')}</td>
                  <td class="dt-mono">{_v(sk,'ID_Posición')}</td>
                  <td class="dt-mono">{_v(sk,'Formato')}</td>
                  <td class="dt-num">{_v(sk,'Cantidad','num')}</td>
                  <td class="dt-num">{_v(sk,'inv Fisico','num')}</td>
                  <td class="dt-num dt-cap">{_v(sk,'Capacidad Utilizada','dec')}</td>
                </tr>"""

            rows += f"""
            <tr class="zona-row" onclick="toggleZona(this,'{rid}')" role="button" tabindex="0"
                aria-expanded="false" title="Click para ver {sku_count} SKUs">
              <td class="td-left td-formato">
                {_zona_chip(zona_str)}
                <span class="expand-hint">▾</span>
              </td>
              {fmt_cell}
              <td class="td-num">{row['Pos_Vol']:,.1f}</td>
              <td class="td-num td-cap">{row['Cap_Real']:,.0f}</td>
              <td class="td-pct {css}">
                <span class="pct-val">{fmt_pct(pv)}</span>
                {gauge_bar(pv, css)}
              </td>
              <td class="td-badge">
                <span class="pill {css}">{semaforo_label(pv)}</span>
                <span class="sku-count-badge">{sku_count} SKUs</span>
              </td>
            </tr>
            <tr class="zona-detail-row" id="zdr-{rid}">
              <td colspan="{n_cols}" class="zona-detail-cell">
                <div class="zona-detail-inner">
                  <div class="zona-detail-content">
                    <div class="zona-detail-search-row">
                      <input type="search" class="detail-search" placeholder="Buscar SKU…"
                             oninput="filterZonaDetail(this,'{rid}')">
                    </div>
                    <div class="detail-table-wrap">
                      <table class="detail-table">
                        <thead>
                          <tr>
                            <th class="th-left">Código</th>
                            <th class="th-left th-desc">Descripción</th>
                            <th>ID Posición</th>
                            <th>Formato</th>
                            <th>Cantidad</th>
                            <th>Inv Físico</th>
                            <th>Cap. Utilizada</th>
                          </tr>
                        </thead>
                        <tbody id="zbody-{rid}">{sku_rows}</tbody>
                      </table>
                    </div>
                  </div>
                </div>
              </td>
            </tr>"""

        sections_html += f"""
        <section class="camara-section" data-bodega="{bodega_label}" aria-label="Cámara {bodega_label}">
          <div class="camara-header">
            <div class="camara-title">
              <span class="camara-icon" aria-hidden="true">{_ICON_WAREHOUSE}</span>
              <h2 class="camara-name">{bodega_label}</h2>
              <span class="badge {badge_cls}" role="status">{badge_txt}</span>
            </div>
            <div class="camara-stats">
              <div class="stat-chip">
                <span class="stat-label">{_ICON_BOX} Capacidad</span>
                <span class="stat-val">{total_cap:,.0f} pos.</span>
              </div>
              <div class="stat-chip {'stat-chip-red' if grp_critico > 0 else ''}">
                <span class="stat-label">{crit_icon} Críticos</span>
                <span class="stat-val {'stat-red' if grp_critico > 0 else 'stat-muted'}">{grp_critico}</span>
              </div>
              <div class="stat-chip">
                <span class="stat-label">{ok_icon} OK</span>
                <span class="stat-val stat-green">{grp_ok}</span>
              </div>
            </div>
          </div>

          <div class="table-wrap" role="region" aria-label="Resumen {bodega_label}" tabindex="0">
            <table>
              <thead>
                <tr>
                  <th scope="col" class="th-left">Zona</th>
                  {'<th scope="col" class="th-left">Formato</th>' if has_fmt else ''}
                  <th scope="col">Pos. Requeridas</th>
                  <th scope="col">Cap. Real</th>
                  <th scope="col">% Ocupación</th>
                  <th scope="col">Estado</th>
                </tr>
              </thead>
              <tbody>{rows}</tbody>
            </table>
          </div>

        </section>"""

    # Global KPI bar HTML
    kpi_cards = f"""
    <div class="global-cards" role="region" aria-label="Resumen global">
      <div class="kpi-card">
        <div class="kpi-icon kpi-icon-blue">{_ICON_BOX}</div>
        <div class="kpi-body">
          <span class="kpi-label">SKUs analizados</span>
          <span class="kpi-val kpi-blue">{total_skus}</span>
        </div>
      </div>
      <div class="kpi-card">
        <div class="kpi-icon {'kpi-icon-amber' if sin_match > 0 else 'kpi-icon-green'}">{_ICON_ALERT if sin_match > 0 else _ICON_CHECK}</div>
        <div class="kpi-body">
          <span class="kpi-label">Sin cámara mapeada</span>
          <span class="kpi-val {'kpi-amber' if sin_match > 0 else 'kpi-green'}">{sin_match}</span>
        </div>
      </div>
      <div class="kpi-card">
        <div class="kpi-icon {'kpi-icon-red' if n_critico > 0 else 'kpi-icon-green'}">{_ICON_ALERT if n_critico > 0 else _ICON_CHECK}</div>
        <div class="kpi-body">
          <span class="kpi-label">Combinaciones críticas</span>
          <span class="kpi-val {'kpi-red' if n_critico > 0 else 'kpi-green'}">{n_critico}</span>
        </div>
      </div>
      <div class="kpi-card">
        <div class="kpi-icon {'kpi-icon-amber' if n_alerta > 0 else 'kpi-icon-green'}">{_ICON_ALERT if n_alerta > 0 else _ICON_CHECK}</div>
        <div class="kpi-body">
          <span class="kpi-label">En alerta (75–85%)</span>
          <span class="kpi-val {'kpi-amber' if n_alerta > 0 else 'kpi-green'}">{n_alerta}</span>
        </div>
      </div>
      <div class="kpi-card">
        <div class="kpi-icon kpi-icon-slate">{_ICON_WAREHOUSE}</div>
        <div class="kpi-body">
          <span class="kpi-label">Cámaras / Bodegas</span>
          <span class="kpi-val kpi-slate">{result['Bodega'].nunique()}</span>
        </div>
      </div>
    </div>"""

    # ── Resumen Ejecutivo ──────────────────────────────────────────────────────
    r_with_cap = result[result["Cap_Real"] > 0]
    total_demand   = r_with_cap["Pos_Vol"].sum()
    total_capacity = r_with_cap["Cap_Real"].sum()
    global_pct     = total_demand / total_capacity if total_capacity > 0 else 0.0
    global_css     = semaforo_css(global_pct)
    global_lbl     = semaforo_label(global_pct)

    # Gauge ring (SVG)
    radius, circ = 48, 301.59
    fill_len = min(global_pct, 1.0) * circ
    gauge_color = {"badge-green": "#22C55E", "badge-yellow": "#F59E0B", "badge-red": "#EF4444"}.get(global_css, "#94A3B8")
    gauge_svg = f"""<svg width="120" height="120" viewBox="0 0 120 120" role="img" aria-label="Ocupación global {global_pct:.1%}">
      <circle cx="60" cy="60" r="{radius}" fill="none" stroke="#E2E8F0" stroke-width="10"/>
      <circle cx="60" cy="60" r="{radius}" fill="none" stroke="{gauge_color}" stroke-width="10"
        stroke-dasharray="{fill_len:.2f} {circ:.2f}"
        stroke-dashoffset="{circ/4:.2f}" stroke-linecap="round"/>
      <text x="60" y="56" text-anchor="middle" font-family="Fira Code,monospace" font-size="17" font-weight="700" fill="{gauge_color}">{global_pct:.1%}</text>
      <text x="60" y="73" text-anchor="middle" font-family="Fira Sans,sans-serif" font-size="9" fill="#64748B">GLOBAL</text>
    </svg>"""

    # Críticas list
    criticas = result[result["Pct_Vol"].fillna(0) > TH_YELLOW].sort_values("Pct_Vol", ascending=False)
    crit_rows = ""
    for _, cr in criticas.iterrows():
        label_parts = [str(cr["Bodega"])]
        if "Zona" in cr.index:   label_parts.append(str(cr["Zona"]))
        if "Formato" in cr.index and pd.notna(cr.get("Formato")): label_parts.append(str(cr["Formato"]))
        pv = cr["Pct_Vol"]
        w  = min(pv * 100, 100)
        crit_rows += f"""
        <div class="ex-crit-row">
          <div class="ex-crit-label">{" / ".join(label_parts)}</div>
          <div class="ex-crit-bar-wrap">
            <div class="ex-crit-bar" style="width:{w:.1f}%;background:#EF4444"></div>
          </div>
          <span class="ex-crit-pct">{pv:.1%}</span>
        </div>"""

    # Bodega summary cards
    bodega_cards = ""
    for b, bg in result.groupby("Bodega", dropna=False):
        bl   = b if pd.notna(b) else "Sin Bodega"
        bc   = bg[bg["Cap_Real"] > 0]
        bd   = bc["Pos_Vol"].sum()
        bcap = bc["Cap_Real"].sum()
        bp   = bd / bcap if bcap > 0 else 0.0
        bcss = semaforo_css(bp)
        bw   = min(bp * 100, 100)
        bcolor = {"badge-green": "#22C55E", "badge-yellow": "#F59E0B", "badge-red": "#EF4444"}.get(bcss, "#94A3B8")
        bodega_cards += f"""
        <div class="ex-bodega-card">
          <div class="ex-bodega-header">
            <span class="ex-bodega-name">{bl}</span>
            <span class="ex-bodega-pct" style="color:{bcolor}">{bp:.1%}</span>
          </div>
          <div class="ex-bar-track">
            <div class="ex-bar-fill" style="width:{bw:.1f}%;background:{bcolor}"></div>
          </div>
          <div class="ex-bodega-sub">{bd:,.0f} / {bcap:,.0f} pos.</div>
        </div>"""

    exec_html = f"""
    <div id="panel-resumen" class="tab-panel" role="tabpanel" aria-labelledby="tab-resumen">

      <div class="ex-grid">

        <!-- Gauge global -->
        <div class="ex-gauge-card">
          <div class="ex-gauge-title">Ocupación Global</div>
          {gauge_svg}
          <div class="ex-gauge-sub">{total_demand:,.0f} pos. usadas de {total_capacity:,.0f}</div>
          <span class="pill {global_css}" style="margin-top:6px">{global_lbl}</span>
        </div>

        <!-- Bodegas -->
        <div class="ex-bodegas-card">
          <div class="ex-section-title">Ocupación por Cámara</div>
          {bodega_cards}
        </div>

        <!-- Críticas -->
        <div class="ex-criticas-card">
          <div class="ex-section-title">
            Zonas Críticas
            <span class="ex-crit-count">{len(criticas)}</span>
          </div>
          {"" if len(criticas) > 0 else '<p class="ex-empty">Sin zonas críticas ✓</p>'}
          {crit_rows}
        </div>

      </div>
    </div>"""

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Ocupación Fábrica</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=Fira+Code:wght@400;500;600&family=Fira+Sans:wght@300;400;500;600;700&display=swap" rel="stylesheet">
  <style>
    /* ── Reset ── */
    *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
    @media (prefers-reduced-motion: reduce) {{
      *, *::before, *::after {{ transition: none !important; animation: none !important; }}
    }}

    /* ── Tokens ── */
    :root {{
      --primary:      #1E40AF;
      --primary-mid:  #2563EB;
      --primary-lt:   #DBEAFE;
      --secondary:    #3B82F6;
      --accent:       #D97706;
      --bg:           #F8FAFC;
      --bg-card:      #FFFFFF;
      --bg-header:    #EFF6FF;
      --border:       #DBEAFE;
      --border-mid:   #BFDBFE;
      --text-main:    #0F172A;
      --text-body:    #1E3A8A;
      --text-muted:   #64748B;
      --text-subtle:  #94A3B8;
      --green:        #16A34A;
      --green-bg:     #F0FDF4;
      --green-border: #BBF7D0;
      --amber:        #D97706;
      --amber-bg:     #FFFBEB;
      --amber-border: #FDE68A;
      --red:          #DC2626;
      --red-bg:       #FEF2F2;
      --red-border:   #FECACA;
      --radius:       10px;
      --radius-sm:    6px;
      --shadow-sm:    0 1px 3px rgba(0,0,0,.07), 0 1px 2px rgba(0,0,0,.05);
      --shadow-md:    0 4px 6px -1px rgba(0,0,0,.08), 0 2px 4px -2px rgba(0,0,0,.05);
    }}

    /* ── Base ── */
    body {{
      font-family: 'Fira Sans', system-ui, sans-serif;
      background: var(--bg);
      color: var(--text-main);
      padding: 32px 28px;
      min-height: 100vh;
      font-size: 14px;
      line-height: 1.6;
    }}

    /* ── Page header ── */
    .page-header {{
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      flex-wrap: wrap;
      gap: 12px;
      margin-bottom: 28px;
      padding-bottom: 20px;
      border-bottom: 2px solid var(--border);
    }}
    .header-left h1 {{
      font-size: 1.6rem;
      font-weight: 700;
      color: var(--primary);
      letter-spacing: -0.02em;
      line-height: 1.2;
    }}
    .header-left .subtitle {{
      color: var(--text-muted);
      font-size: 0.82rem;
      margin-top: 4px;
      font-weight: 400;
    }}
    .header-badge {{
      background: var(--primary-lt);
      color: var(--primary);
      font-size: 0.72rem;
      font-weight: 600;
      padding: 5px 12px;
      border-radius: 20px;
      border: 1px solid var(--border-mid);
      white-space: nowrap;
      align-self: center;
    }}

    /* ── KPI cards ── */
    .global-cards {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
      gap: 8px;
      margin-bottom: 18px;
    }}
    .kpi-card {{
      background: var(--bg-card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      padding: 10px 12px;
      display: flex;
      align-items: center;
      gap: 10px;
      box-shadow: var(--shadow-sm);
      transition: box-shadow 150ms ease, border-color 150ms ease;
    }}
    .kpi-card:hover {{
      box-shadow: var(--shadow-md);
      border-color: var(--border-mid);
    }}
    .kpi-icon {{
      width: 28px; height: 28px;
      border-radius: 6px;
      display: flex; align-items: center; justify-content: center;
      flex-shrink: 0;
    }}
    .kpi-icon-blue  {{ background: #DBEAFE; color: var(--primary-mid); }}
    .kpi-icon-red   {{ background: #FEE2E2; color: var(--red); }}
    .kpi-icon-amber {{ background: #FEF3C7; color: var(--amber); }}
    .kpi-icon-green {{ background: #DCFCE7; color: var(--green); }}
    .kpi-icon-slate {{ background: #F1F5F9; color: var(--text-muted); }}
    .kpi-body {{ display: flex; flex-direction: column; gap: 0; }}
    .kpi-label {{ font-size: 0.6rem; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: .07em; }}
    .kpi-val {{
      font-family: 'Fira Code', monospace;
      font-size: 1.3rem; font-weight: 700; line-height: 1.15;
    }}
    .kpi-blue  {{ color: var(--primary-mid); }}
    .kpi-red   {{ color: var(--red); }}
    .kpi-amber {{ color: var(--amber); }}
    .kpi-green {{ color: var(--green); }}
    .kpi-slate {{ color: var(--text-muted); }}

    /* ── Leyenda ── */
    .legend {{
      display: flex; gap: 16px; flex-wrap: wrap;
      align-items: center; margin-bottom: 24px;
    }}
    .legend-title {{
      font-size: 0.72rem; font-weight: 600; color: var(--text-muted);
      text-transform: uppercase; letter-spacing: .06em;
    }}
    .legend-item {{
      display: flex; align-items: center; gap: 6px;
      font-size: 0.78rem; color: var(--text-body); font-weight: 400;
    }}
    .ldot {{
      width: 10px; height: 10px; border-radius: 3px; flex-shrink: 0;
    }}
    .ldot-green  {{ background: var(--green); }}
    .ldot-amber  {{ background: var(--amber); }}
    .ldot-red    {{ background: var(--red); }}
    .ldot-gray   {{ background: #CBD5E1; }}

    /* ── Sección cámara ── */
    .camara-section {{
      background: var(--bg-card);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      margin-bottom: 24px;
      overflow: hidden;
      box-shadow: var(--shadow-sm);
    }}
    .camara-header {{
      display: flex;
      align-items: center;
      justify-content: space-between;
      flex-wrap: wrap;
      gap: 12px;
      padding: 16px 20px;
      background: var(--bg-header);
      border-bottom: 1px solid var(--border);
    }}
    .camara-title {{
      display: flex; align-items: center; gap: 10px;
    }}
    .camara-icon {{
      color: var(--primary-mid);
      display: flex; align-items: center;
    }}
    .camara-name {{
      font-size: 1rem; font-weight: 700;
      color: var(--primary); letter-spacing: -0.01em;
    }}

    /* Badges */
    .badge {{
      font-size: 0.65rem; font-weight: 700; letter-spacing: .07em;
      padding: 3px 10px; border-radius: 20px; text-transform: uppercase;
      border: 1px solid;
    }}
    .badge-green  {{ background: var(--green-bg);  color: var(--green);  border-color: var(--green-border); }}
    .badge-yellow {{ background: var(--amber-bg);  color: var(--amber);  border-color: var(--amber-border); }}
    .badge-red    {{ background: var(--red-bg);    color: var(--red);    border-color: var(--red-border); }}
    .badge-gray   {{ background: #F1F5F9; color: var(--text-muted); border-color: #E2E8F0; }}

    /* Stats chips */
    .camara-stats {{ display: flex; gap: 16px; flex-wrap: wrap; align-items: center; }}
    .stat-chip {{
      display: flex; flex-direction: column; gap: 1px;
      padding: 6px 12px; border-radius: 8px;
      background: var(--bg-card); border: 1px solid var(--border);
      min-width: 80px;
    }}
    .stat-chip-red {{ background: var(--red-bg); border-color: var(--red-border); }}
    .stat-label {{
      font-size: 0.65rem; color: var(--text-muted); text-transform: uppercase;
      letter-spacing: .05em; font-weight: 500;
      display: flex; align-items: center; gap: 4px;
    }}
    .stat-icon {{ display: inline-flex; align-items: center; }}
    .icon-red   {{ color: var(--red); }}
    .icon-green {{ color: var(--green); }}
    .stat-val {{
      font-family: 'Fira Code', monospace;
      font-size: 1.1rem; font-weight: 600; color: var(--text-main);
    }}
    .stat-red   {{ color: var(--red); }}
    .stat-green {{ color: var(--green); }}
    .stat-muted {{ color: var(--text-muted); }}

    /* ── Tabla ── */
    .table-wrap {{ overflow-x: auto; }}
    table {{
      width: 100%; border-collapse: collapse;
      font-size: 0.82rem;
    }}
    thead {{
      background: #F1F5F9;
      position: sticky; top: 0; z-index: 1;
      border-bottom: 2px solid var(--border-mid);
    }}
    thead th {{
      padding: 10px 14px; text-align: center;
      font-weight: 600; color: var(--text-body);
      white-space: nowrap; font-size: 0.75rem;
      text-transform: uppercase; letter-spacing: .05em;
    }}
    .th-left {{ text-align: left !important; }}

    tbody tr {{
      border-bottom: 1px solid #F1F5F9;
      transition: background 150ms ease;
    }}
    tbody tr:hover {{ background: var(--bg-header); }}
    tbody td {{ padding: 9px 14px; color: var(--text-muted); vertical-align: middle; }}

    .td-left   {{ text-align: left; }}
    .td-formato {{ font-weight: 600; color: var(--text-main); }}
    .td-nivel  {{ color: var(--text-body); }}
    .td-num    {{
      text-align: right; font-family: 'Fira Code', monospace;
      font-size: 0.8rem; color: var(--text-body); font-variant-numeric: tabular-nums;
    }}
    .td-cap    {{ font-weight: 600; color: var(--primary); }}
    .td-pct    {{ text-align: center; padding: 6px 10px; }}
    .td-badge  {{ text-align: center; }}

    /* Gauge bar */
    .pct-val {{
      font-family: 'Fira Code', monospace;
      font-size: 0.82rem; font-weight: 600;
      display: block; margin-bottom: 3px;
      font-variant-numeric: tabular-nums;
    }}
    .bar-wrap {{
      height: 4px; border-radius: 2px;
      background: #E2E8F0; overflow: hidden; width: 100%; min-width: 60px;
    }}
    .bar {{ height: 100%; border-radius: 2px; transition: width 300ms ease; }}

    /* Cell semáforo (pct) */
    td.cell-green  .pct-val {{ color: var(--green); }}
    td.cell-green  .bar      {{ background: var(--green); }}
    td.cell-yellow .pct-val  {{ color: var(--amber); }}
    td.cell-yellow .bar      {{ background: var(--amber); }}
    td.cell-red    .pct-val  {{ color: var(--red); }}
    td.cell-red    .bar      {{ background: var(--red); }}
    td.cell-nocap  .pct-val  {{ color: var(--text-subtle); font-style: italic; }}

    /* Pills (estado) */
    .pill {{
      display: inline-block;
      font-size: 0.65rem; font-weight: 700; letter-spacing: .07em;
      padding: 3px 8px; border-radius: 20px; text-transform: uppercase;
      border: 1px solid;
    }}
    .pill.cell-green  {{ background: var(--green-bg);  color: var(--green);  border-color: var(--green-border); }}
    .pill.cell-yellow {{ background: var(--amber-bg);  color: var(--amber);  border-color: var(--amber-border); }}
    .pill.cell-red    {{ background: var(--red-bg);    color: var(--red);    border-color: var(--red-border); }}
    .pill.cell-nocap  {{ background: #F8FAFC; color: var(--text-subtle); border-color: #E2E8F0; }}

    /* ── Footer ── */
    .footer {{
      margin-top: 28px; padding-top: 16px;
      border-top: 1px solid var(--border);
      font-size: 0.72rem; color: var(--text-subtle);
      display: flex; gap: 8px; align-items: flex-start; flex-wrap: wrap;
    }}
    .footer-icon {{ color: var(--text-subtle); flex-shrink: 0; margin-top: 1px; }}

    /* ── Responsive ── */
    @media (max-width: 768px) {{
      body {{ padding: 16px 12px; }}
      .header-left h1 {{ font-size: 1.3rem; }}
      .camara-header {{ flex-direction: column; align-items: flex-start; }}
      .global-cards {{ grid-template-columns: repeat(2, 1fr); }}
    }}
    @media (max-width: 480px) {{
      .global-cards {{ grid-template-columns: 1fr; }}
      .legend {{ gap: 10px; }}
    }}

    /* ── Filter bar ── */
    .filter-bar {{
      display: flex; align-items: center; gap: 12px;
      padding: 12px 28px 0;
      flex-wrap: wrap;
    }}
    .filter-label {{ font-size: .72rem; font-weight: 700; color: #64748B; text-transform: uppercase; letter-spacing: .06em; white-space: nowrap; }}
    .filter-chips  {{ display: flex; gap: 6px; flex-wrap: wrap; }}
    .filter-chip {{
      font-size: .75rem; font-weight: 600; padding: 4px 12px;
      border: 1.5px solid #E2E8F0; border-radius: 20px;
      background: #F8FAFC; color: #475569;
      cursor: pointer; transition: all 120ms;
    }}
    .filter-chip:hover  {{ background: #EFF6FF; border-color: #93C5FD; color: #1E40AF; }}
    .filter-chip.active {{ background: #1E40AF; border-color: #1E40AF; color: #fff; }}
    .filter-chip-all.active {{ background: #334155; border-color: #334155; color: #fff; }}

    /* ── Clickable zona row ── */
    .zona-row {{
      cursor: pointer; user-select: none;
      transition: background 80ms ease-out;
      touch-action: manipulation;
    }}
    .zona-row:hover  {{ background: #F0F9FF !important; }}
    .zona-row:focus-visible {{
      outline: 2px solid #3B82F6; outline-offset: -2px;
    }}
    .zona-row[aria-expanded="true"] {{ background: #EFF6FF; }}
    .zona-row[aria-expanded="true"] .expand-hint {{ transform: rotate(180deg); }}
    .expand-hint {{
      font-size: .65rem; color: #94A3B8; margin-left: 5px;
      display: inline-block;
      transition: transform 200ms ease-out;
    }}
    .sku-count-badge {{
      display: inline-block; margin-left: 6px;
      font-size: .62rem; font-weight: 600; padding: 1px 6px;
      background: #F1F5F9; border-radius: 10px; color: #64748B;
      vertical-align: middle;
    }}

    /* ── Inline zona detail (animated) ── */
    .zona-detail-row {{ background: #F8FAFC; }}
    .zona-detail-cell {{ padding: 0 !important; border-top: none !important; }}
    .zona-detail-inner {{
      display: grid;
      grid-template-rows: 0fr;
      transition: grid-template-rows 220ms ease-out;
    }}
    .zona-detail-inner > .zona-detail-content {{ overflow: hidden; }}
    .zona-detail-row.open .zona-detail-inner {{
      grid-template-rows: 1fr;
    }}
    .zona-detail-search-row {{
      padding: 8px 12px 6px;
      border-bottom: 1px solid #E2E8F0;
      background: #F1F5F9;
    }}
    .detail-search {{
      font-size: .78rem; padding: 5px 12px;
      border: 1.5px solid #E2E8F0; border-radius: 8px;
      outline: none; color: #1E293B; width: 240px;
      transition: border-color 150ms;
    }}
    .detail-search:focus {{ border-color: #3B82F6; box-shadow: 0 0 0 3px rgba(59,130,246,.15); }}

    /* ── Detail table ── */
    .detail-table-wrap {{ max-height: 320px; overflow-y: auto; }}
    .detail-table {{ width: 100%; border-collapse: collapse; font-size: .78rem; }}
    .detail-table thead th {{
      position: sticky; top: 0; z-index: 1;
      background: #EFF6FF; padding: 7px 10px;
      text-align: left; font-size: .68rem; font-weight: 700;
      color: #1E40AF; text-transform: uppercase; letter-spacing: .05em;
      border-bottom: 1px solid #BFDBFE;
    }}
    .detail-table tbody tr {{ border-bottom: 1px solid #F1F5F9; transition: background 80ms; }}
    .detail-table tbody tr:hover {{ background: #EFF6FF; }}
    .detail-table tbody tr.hidden-row {{ display: none; }}
    .dt-row td {{ padding: 5px 10px; vertical-align: middle; }}
    .dt-code {{ font-family: 'Fira Code', monospace; font-weight: 600; color: #1E40AF; white-space: nowrap; }}
    .dt-desc {{ max-width: 220px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; color: #1E293B; }}
    .dt-mono {{ font-family: 'Fira Code', monospace; font-size: .74rem; color: #475569; }}
    .dt-num  {{ text-align: right; font-variant-numeric: tabular-nums; color: #0F172A; }}
    .dt-cap  {{ font-weight: 700; color: #1E40AF; }}

    /* zona chip in detail */
    .zona-chip  {{ display: inline-block; font-size: .63rem; font-weight: 700; padding: 2px 6px; border-radius: 8px; vertical-align: middle; }}
    .zona-label {{ font-size: .76rem; margin-left: 5px; color: #374151; }}

    /* ── PDF export button ── */
    .pdf-btn {{
      display: inline-flex; align-items: center; gap: 6px;
      font-family: 'Fira Sans', sans-serif;
      font-size: .75rem; font-weight: 600; padding: 6px 14px;
      background: #1E40AF; border: none;
      border-radius: 8px; color: #fff; cursor: pointer;
      transition: background 150ms, box-shadow 150ms;
      white-space: nowrap; box-shadow: 0 1px 3px rgba(30,64,175,.3);
    }}
    .pdf-btn:hover {{ background: #1D4ED8; box-shadow: 0 3px 8px rgba(30,64,175,.4); }}
    .pdf-btn:focus-visible {{ outline: 2px solid #93C5FD; outline-offset: 2px; }}

    /* ── Tab bar ── */
    .tab-bar {{
      display: flex; gap: 4px;
      border-bottom: 2px solid var(--border);
      margin-bottom: 20px;
    }}
    .tab-btn {{
      display: inline-flex; align-items: center; gap: 6px;
      font-family: 'Fira Sans', sans-serif;
      font-size: .8rem; font-weight: 600;
      padding: 9px 18px; border: none; background: none;
      color: var(--text-muted); cursor: pointer;
      border-bottom: 2px solid transparent; margin-bottom: -2px;
      transition: color 150ms, border-color 150ms;
      border-radius: 6px 6px 0 0;
    }}
    .tab-btn:hover   {{ color: var(--primary-mid); background: #F0F9FF; }}
    .tab-btn.active  {{ color: var(--primary-mid); border-bottom-color: var(--primary-mid); background: #F8FAFC; }}
    .tab-btn:focus-visible {{ outline: 2px solid #3B82F6; outline-offset: 2px; }}
    .tab-panel[hidden] {{ display: none; }}

    /* ── Executive summary layout ── */
    .ex-grid {{
      display: grid;
      grid-template-columns: 180px 1fr 1fr;
      gap: 16px;
      margin-bottom: 24px;
    }}
    @media (max-width: 900px) {{
      .ex-grid {{ grid-template-columns: 1fr 1fr; }}
      .ex-gauge-card {{ grid-column: 1 / -1; }}
    }}
    @media (max-width: 560px) {{
      .ex-grid {{ grid-template-columns: 1fr; }}
    }}

    .ex-gauge-card, .ex-bodegas-card, .ex-criticas-card {{
      background: var(--bg-card); border: 1px solid var(--border);
      border-radius: var(--radius); padding: 16px 18px;
      box-shadow: var(--shadow-sm);
    }}
    .ex-gauge-card {{
      display: flex; flex-direction: column; align-items: center;
      gap: 6px; text-align: center;
    }}
    .ex-gauge-title {{
      font-size: .7rem; font-weight: 700; text-transform: uppercase;
      letter-spacing: .07em; color: var(--text-muted);
    }}
    .ex-gauge-sub {{ font-size: .72rem; color: var(--text-muted); }}

    .ex-section-title {{
      font-size: .72rem; font-weight: 700; text-transform: uppercase;
      letter-spacing: .07em; color: var(--text-muted);
      margin-bottom: 12px; display: flex; align-items: center; gap: 8px;
    }}
    .ex-crit-count {{
      background: #FEE2E2; color: #DC2626;
      font-size: .65rem; font-weight: 700; padding: 1px 7px;
      border-radius: 10px;
    }}
    .ex-empty {{ font-size: .82rem; color: #22C55E; font-weight: 600; padding: 8px 0; }}

    /* Bodega cards */
    .ex-bodega-card {{ margin-bottom: 12px; }}
    .ex-bodega-card:last-child {{ margin-bottom: 0; }}
    .ex-bodega-header {{
      display: flex; justify-content: space-between; align-items: baseline;
      margin-bottom: 4px;
    }}
    .ex-bodega-name {{
      font-size: .78rem; font-weight: 600; color: var(--text-main);
      white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 75%;
    }}
    .ex-bodega-pct {{
      font-family: 'Fira Code', monospace; font-size: .8rem; font-weight: 700;
    }}
    .ex-bar-track {{
      height: 7px; background: #F1F5F9; border-radius: 4px; overflow: hidden;
    }}
    .ex-bar-fill {{
      height: 100%; border-radius: 4px;
      transition: width 400ms ease-out;
    }}
    .ex-bodega-sub {{
      font-size: .65rem; color: var(--text-muted); margin-top: 3px;
      font-family: 'Fira Code', monospace;
    }}

    /* Críticas list */
    .ex-crit-row {{
      display: flex; align-items: center; gap: 8px; margin-bottom: 8px;
    }}
    .ex-crit-label {{
      font-size: .72rem; color: var(--text-main); white-space: nowrap;
      overflow: hidden; text-overflow: ellipsis; flex: 0 0 42%;
    }}
    .ex-crit-bar-wrap {{
      flex: 1; height: 7px; background: #FEE2E2; border-radius: 4px; overflow: hidden;
    }}
    .ex-crit-bar {{
      height: 100%; border-radius: 4px;
      transition: width 400ms ease-out;
    }}
    .ex-crit-pct {{
      font-family: 'Fira Code', monospace; font-size: .72rem; font-weight: 700;
      color: #DC2626; flex: 0 0 44px; text-align: right;
    }}

    /* ── Focus rings ── */
    .filter-chip:focus-visible {{
      outline: 2px solid #3B82F6; outline-offset: 2px;
    }}
    .up-btn:focus-visible {{
      outline: 2px solid #1E40AF; outline-offset: 2px;
    }}

    /* ── Reduced motion ── */
    @media (prefers-reduced-motion: reduce) {{
      *, *::before, *::after {{
        transition-duration: .01ms !important;
        animation-duration: .01ms !important;
      }}
    }}

    /* ── Print / PDF ── */
    @media print {{
      body {{ background: white; padding: 16px; font-size: 11pt; }}
      .tab-bar, .filter-bar, .pdf-btn, .zona-detail-row {{ display: none !important; }}
      .tab-panel[hidden] {{ display: block !important; }}
      #panel-detalle {{ display: none !important; }}
      .camara-section {{ break-inside: avoid; }}
      .ex-grid {{ grid-template-columns: 160px 1fr 1fr; }}
      thead {{ position: static; }}
      .page-header {{ border-bottom: 1px solid #ccc; margin-bottom: 12px; }}
    }}
  </style>
</head>
<body>

  <div class="page-header">
    <div class="header-left">
      <h1>Ocupación Fábrica</h1>
      <p class="subtitle">Fuente: <em>{Path(INPUT_FILE).name}</em>{(" &nbsp;·&nbsp; " + LABEL) if LABEL else ""}</p>
    </div>
    <button class="pdf-btn" onclick="exportPDF()" title="Exportar Resumen Ejecutivo como PDF">
      <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>
      Exportar PDF
    </button>
  </div>

  {kpi_cards}

  <!-- Tab bar -->
  <div class="tab-bar" role="tablist" aria-label="Vistas del reporte">
    <button class="tab-btn active" id="tab-resumen" role="tab" aria-selected="true"
            aria-controls="panel-resumen" onclick="switchTab('resumen')">
      <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><rect x="3" y="3" width="7" height="7"/><rect x="14" y="3" width="7" height="7"/><rect x="3" y="14" width="7" height="7"/><rect x="14" y="14" width="7" height="7"/></svg>
      Resumen Ejecutivo
    </button>
    <button class="tab-btn" id="tab-detalle" role="tab" aria-selected="false"
            aria-controls="panel-detalle" onclick="switchTab('detalle')">
      <svg width="13" height="13" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><line x1="8" y1="6" x2="21" y2="6"/><line x1="8" y1="12" x2="21" y2="12"/><line x1="8" y1="18" x2="21" y2="18"/><line x1="3" y1="6" x2="3.01" y2="6"/><line x1="3" y1="12" x2="3.01" y2="12"/><line x1="3" y1="18" x2="3.01" y2="18"/></svg>
      Detalle por Cámara
    </button>
  </div>

  {exec_html}

  <div id="panel-detalle" class="tab-panel" role="tabpanel" aria-labelledby="tab-detalle" hidden>
    <div class="legend" role="note" aria-label="Leyenda semáforos">
      <span class="legend-title">Semáforo:</span>
      <div class="legend-item"><div class="ldot ldot-green" aria-hidden="true"></div>&lt; {TH_GREEN:.0%} — OK</div>
      <div class="legend-item"><div class="ldot ldot-amber" aria-hidden="true"></div>{TH_GREEN:.0%}–{TH_YELLOW:.0%} — Alerta</div>
      <div class="legend-item"><div class="ldot ldot-red"  aria-hidden="true"></div>&gt; {TH_YELLOW:.0%} — Crítico</div>
      <div class="legend-item"><div class="ldot ldot-gray" aria-hidden="true"></div>Sin capacidad mapeada en inf</div>
    </div>

    <div class="filter-bar" role="navigation" aria-label="Filtro por cámara">
      <span class="filter-label">Cámara:</span>
      <div class="filter-chips">
        <button class="filter-chip filter-chip-all active" onclick="filterCam(this,null)">Todas</button>
        {filter_chips}
      </div>
    </div>

    <main id="main-sections">
      {sections_html}
    </main>
  </div>

  <footer class="footer">
    <span class="footer-icon" aria-hidden="true">{_ICON_INFO}</span>
    <span>
      Pos. requeridas = Capacidad Utilizada por SKU desde BASE &nbsp;·&nbsp;
      Umbrales: Verde &lt;{TH_GREEN:.0%} · Amarillo {TH_GREEN:.0%}–{TH_YELLOW:.0%} · Rojo &gt;{TH_YELLOW:.0%}
    </span>
  </footer>

<script>
  /* ── Tab switching ── */
  function switchTab(name) {{
    ['resumen','detalle'].forEach(t => {{
      const btn   = document.getElementById('tab-' + t);
      const panel = document.getElementById('panel-' + t);
      const active = t === name;
      btn.classList.toggle('active', active);
      btn.setAttribute('aria-selected', active);
      panel.hidden = !active;
    }});
  }}

  /* ── Camera filter (multi-select) ── */
  function filterCam(btn, bodega) {{
    const allBtn = document.querySelector('.filter-chip-all');
    if (!bodega) {{
      // "Todas" → deselect all specific chips, show all
      document.querySelectorAll('.filter-chip:not(.filter-chip-all)').forEach(b => b.classList.remove('active'));
      allBtn.classList.add('active');
      document.querySelectorAll('#main-sections .camara-section').forEach(s => s.hidden = false);
      return;
    }}
    btn.classList.toggle('active');
    const selected = [...document.querySelectorAll('.filter-chip:not(.filter-chip-all).active')]
                       .map(b => b.dataset.bodega);
    if (selected.length === 0) {{
      allBtn.classList.add('active');
      document.querySelectorAll('#main-sections .camara-section').forEach(s => s.hidden = false);
    }} else {{
      allBtn.classList.remove('active');
      document.querySelectorAll('#main-sections .camara-section').forEach(s => {{
        s.hidden = !selected.includes(s.dataset.bodega);
      }});
    }}
  }}

  /* ── Zona row expand (animated via CSS grid-template-rows) ── */
  function toggleZona(tr, id) {{
    const detail = document.getElementById('zdr-' + id);
    const open   = !detail.classList.contains('open');
    detail.classList.toggle('open', open);
    tr.setAttribute('aria-expanded', open);
    if (!open) {{
      const inp = detail.querySelector('.detail-search');
      if (inp) {{ inp.value = ''; filterZonaDetail(inp, id); }}
    }}
  }}

  /* ── Keyboard support for zona rows ── */
  document.addEventListener('keydown', e => {{
    if ((e.key === 'Enter' || e.key === ' ') && e.target.classList.contains('zona-row')) {{
      e.preventDefault();
      const id = e.target.getAttribute('onclick').match(/'([^']+)'\)$/)[1];
      toggleZona(e.target, id);
    }}
  }});

  /* ── Detail search per zona ── */
  function filterZonaDetail(inp, id) {{
    const q = inp.value.toLowerCase();
    document.querySelectorAll('#zbody-' + id + ' .dt-row').forEach(row => {{
      row.classList.toggle('hidden-row', q.length > 0 && !row.textContent.toLowerCase().includes(q));
    }});
  }}

  /* ── PDF export (prints Resumen Ejecutivo only) ── */
  function exportPDF() {{
    switchTab('resumen');
    setTimeout(() => window.print(), 120);
  }}
</script>

</body>
</html>"""

    Path(path).write_text(html, encoding="utf-8")
    print(f"  HTML  → {path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    global INPUT_FILE, CAPACIDADES_FILE, LABEL, OUTPUT_XLSX, OUTPUT_HTML, OUTPUT_DIR

    parser = argparse.ArgumentParser(description="Análisis ocupación logística")
    parser.add_argument("--base",       default=None, help="Archivo de stock físico (.xlsx)")
    parser.add_argument("--cap",        default=None, help="Tabla de capacidades (template_capacidades.xlsx)")
    parser.add_argument("--label",      default="",   help="Etiqueta para el reporte (ej: 'Promedio Semanal')")
    parser.add_argument("--output-dir", default=None, help="Directorio de salida para los archivos generados")
    args = parser.parse_args()

    if args.base:       INPUT_FILE       = args.base
    if args.cap:        CAPACIDADES_FILE = args.cap
    if args.label:      LABEL            = args.label
    if args.output_dir:
        OUTPUT_DIR  = args.output_dir
        OUTPUT_XLSX = str(Path(OUTPUT_DIR) / "reporte_ocupacion_sep.xlsx")
        OUTPUT_HTML = str(Path(OUTPUT_DIR) / "reporte_ocupacion_sep.html")

    print(f"[CONFIG] base={INPUT_FILE} | cap={CAPACIDADES_FILE or 'inf'} | label='{LABEL}' | out={OUTPUT_DIR or '.'}")

    print(f"[1/3] Leyendo stock: {INPUT_FILE}")
    xls = pd.ExcelFile(INPUT_FILE)
    stock, _ = load_sku_master(xls)
    sin_bodega = stock["Bodega"].isna().sum()
    if sin_bodega:
        print(f"      ⚠  {sin_bodega} SKUs sin bodega")

    print("[2/3] Capacidades...")
    capacity = load_capacity(xls)

    print("[3/3] Consolidando...")
    result = consolidate(stock, capacity)

    print("\n── RESUMEN ──────────────────────────────────────────────────────")
    for _, r in result.iterrows():
        pv_str = f"{r['Pct_Vol']:.1%}" if pd.notna(r["Pct_Vol"]) else "—"
        estado = semaforo_label(r["Pct_Vol"])
        print(f"  {str(r['Bodega']):<25} {str(r['Zona']):<8} {pv_str:>7}  [{estado}]")

    print("\n[OUTPUT]")
    write_excel(result, OUTPUT_XLSX)
    write_html(result, stock, OUTPUT_HTML)
    print("✅ Listo.")


if __name__ == "__main__":
    main()
