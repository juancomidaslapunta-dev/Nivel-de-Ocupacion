"""
Genera template_capacidades.xlsx con estructura simplificada:
  Bodega | Zona | ID_Posición | Num_Slots | Cap. Nominal Bandejas | Cap. Nominal Racks

Pre-llena con datos reales. Usuario solo agrega/edita filas.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = "template_capacidades.xlsx"

# ── Datos iniciales (modificar aquí o directamente en el Excel) ───────────────
ROWS = [
    # Bodega                  Zona   ID_Pos     Slots  Cap_Band  Cap_Rack
    ("Congelado salado 5",   "Rack", "N1-N3",   42,    None,     42  ),
    ("Congelado salado 5",   "Rack", "N4",       14,    None,     14  ),
    ("Congelado salado 5",   "Piso", "P1-P4",    4,     72,       None),
    ("Mantencion salado 4",  "Rack", "N1-N3",    9,     None,     9   ),
    ("Mantencion salado 4",  "Rack", "N4",        3,     None,     3   ),
    ("Mantencion salado 4",  "Piso", "P5-P15",   12,    120,      None),
    ("Congelado Reefer 7",   "Piso", "P20-P40",  23,    345,      None),
    ("Congelado Reefer 10",  "Piso", "P50-P60",  23,    345,      None),
    ("Congelado Reefer 11",  "Piso", "P70-P90",  23,    345,      None),
]

# ── Estilos ───────────────────────────────────────────────────────────────────
C = {
    "hdr_bg":    "1E40AF",
    "hdr_fg":    "FFFFFF",
    "rack_bg":   "EFF6FF",
    "piso_bg":   "F0FDF4",
    "band_bg":   "FFFBEB",  # Cap Bandejas editable
    "racks_bg":  "EFF6FF",  # Cap Racks editable
    "alt_rack":  "DBEAFE",
    "alt_piso":  "DCFCE7",
    "border":    "BFDBFE",
    "muted":     "94A3B8",
    "primary":   "1E40AF",
    "green":     "16A34A",
    "none_fg":   "CBD5E1",  # color para "—"
}

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="0F172A", sz=9, italic=False):
    return Font(bold=bold, color=color, size=sz,
                name="Segoe UI", italic=italic)
def border(color=C["border"]):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="center", v="center", wrap=False, indent=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

HEADERS = [
    "Bodega",
    "Zona",
    "ID_Posición",
    "Num_Slots",
    "Capacidad Nominal\nBandejas",
    "Capacidad Nominal\nRacks",
]
WIDTHS = [24, 10, 14, 12, 22, 20]

def write_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Capacidades"

    # ── Título ──
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "CAPACIDADES POR BODEGA — EDITABLE"
    t.fill      = fill(C["hdr_bg"])
    t.font      = Font(bold=True, color=C["hdr_fg"], size=11, name="Segoe UI")
    t.alignment = align()
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = (
        "Agregar filas según necesidad · "
        "Zona: Rack = posiciones de rack  |  Piso = posiciones de bandeja  · "
        "Dejar en blanco lo que no aplique"
    )
    sub.font      = font(color=C["muted"], italic=True, sz=8)
    sub.alignment = align(h="left", indent=1)
    sub.fill      = fill("F8FAFC")
    ws.row_dimensions[2].height = 15

    # ── Headers ──
    for ci, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        c = ws.cell(3, ci, h)
        c.fill      = fill(C["hdr_bg"])
        c.font      = font(bold=True, color=C["hdr_fg"], sz=9)
        c.alignment = align(wrap=True)
        c.border    = border("1E40AF")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 36

    # ── Validaciones ──
    dv_zona = DataValidation(
        type="list", formula1='"Rack,Piso"',
        allow_blank=False, showDropDown=False,
        showErrorMessage=True, error="Usar Rack o Piso",
    )
    ws.add_data_validation(dv_zona)

    # ── Datos ──
    prev_bodega = None
    shade = False

    for ri, (bodega, zona, id_pos, slots, cap_band, cap_rack) in enumerate(ROWS, 4):
        if bodega != prev_bodega:
            shade = not shade
            prev_bodega = bodega

        is_rack = zona == "Rack"
        row_fill_base = C["alt_rack"] if (shade and is_rack) else \
                        C["rack_bg"]  if is_rack else \
                        C["alt_piso"] if shade else C["piso_bg"]

        vals = [bodega, zona, id_pos, slots, cap_band, cap_rack]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(ri, ci)
            c.border = border()
            c.alignment = align(h="center" if ci > 1 else "left",
                                indent=1 if ci == 1 else 0)

            if val is None:
                c.value     = "—"
                c.font      = font(color=C["none_fg"], italic=True)
                c.fill      = fill("F1F5F9")
            else:
                c.value = val
                if ci == 1:   # Bodega
                    c.font = font(bold=True, color=C["primary"])
                    c.fill = fill(row_fill_base)
                elif ci == 2:  # Zona badge
                    c.font = font(bold=True,
                                  color="1D4ED8" if is_rack else "15803D", sz=9)
                    c.fill = fill("DBEAFE" if is_rack else "DCFCE7")
                elif ci == 3:  # ID_Posicion
                    c.font = Font(name="Consolas", size=9, color="374151")
                    c.fill = fill(row_fill_base)
                elif ci == 4:  # Num_Slots
                    c.font = Font(name="Consolas", bold=True, size=9, color="0F172A")
                    c.fill = fill(row_fill_base)
                    c.alignment = align(h="right")
                elif ci == 5:  # Cap Bandejas
                    c.font = Font(name="Consolas", bold=True, size=10,
                                  color=C["green"] if not is_rack else "CBD5E1")
                    c.fill = fill(C["piso_bg"] if not is_rack else "F8FAFC")
                    c.alignment = align(h="right")
                elif ci == 6:  # Cap Racks
                    c.font = Font(name="Consolas", bold=True, size=10,
                                  color="1D4ED8" if is_rack else "CBD5E1")
                    c.fill = fill("EFF6FF" if is_rack else "F8FAFC")
                    c.alignment = align(h="right")

            dv_zona.sqref += f"B{ri}"

        ws.row_dimensions[ri].height = 19

    # ── Fila de total ──
    last_row = 3 + len(ROWS)
    total_row = last_row + 1
    ws.merge_cells(f"A{total_row}:D{total_row}")
    t = ws.cell(total_row, 1, "TOTALES")
    t.font      = font(bold=True, color=C["primary"])
    t.fill      = fill("EFF6FF")
    t.alignment = align(h="right")
    t.border    = border("BFDBFE")

    for ci, col_letter in enumerate(["E", "F"], 5):
        c = ws.cell(total_row, ci,
                    f"=SUMIF(B4:B{last_row},\"{'Piso' if ci==5 else 'Rack'}\","
                    f"{col_letter}4:{col_letter}{last_row})")
        c.font      = Font(name="Consolas", bold=True, size=10,
                           color=C["green"] if ci == 5 else "1D4ED8")
        c.fill      = fill("F0FDF4" if ci == 5 else "EFF6FF")
        c.alignment = align(h="right")
        c.border    = border("BFDBFE")
    ws.row_dimensions[total_row].height = 20

    # ── Freeze + filtros ──
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{last_row}"

    # ── Hoja instrucciones ──
    wi = wb.create_sheet("Instrucciones")
    wi.column_dimensions["A"].width = 80
    lines = [
        ("INSTRUCCIONES", True),
        ("", False),
        ("1. ESTRUCTURA", True),
        ("   Bodega     → Nombre de la bodega/cámara (ej: 'Congelado salado 5')", False),
        ("   Zona       → 'Rack' para estructuras de rack | 'Piso' para posiciones de bandeja en piso", False),
        ("   ID_Posición → Identificador del bloque (ej: N1-N3, P1-P4)", False),
        ("   Num_Slots  → Cantidad de posiciones físicas en ese bloque", False),
        ("   Cap. Bandejas → Capacidad total en bandejas (solo Zona=Piso)", False),
        ("   Cap. Racks    → Capacidad total en posiciones rack (solo Zona=Rack)", False),
        ("", False),
        ("2. CÓMO AGREGAR UNA BODEGA NUEVA", True),
        ("   Agregar filas al final con el mismo formato.", False),
        ("   El script leerá todas las filas a partir de la fila 4.", False),
        ("", False),
        ("3. MAPEO A NOMBRES BASE", True),
        ("   El script normaliza automáticamente:", False),
        ("   'Congelado salado 5'  → 'Congelado salado'", False),
        ("   'Mantencion salado 4' → 'Mantencion salado'", False),
        ("   Para otras bodegas, el nombre debe coincidir exactamente con BASE.", False),
        ("", False),
        ("4. LÓGICA DE OCUPACIÓN", True),
        ("   Zona=Rack  → Demanda en posiciones de rack vs Cap. Racks", False),
        ("   Zona=Piso  → Demanda en bandejas (todas las familias) vs Cap. Bandejas", False),
        ("", False),
        ("5. GUARDAR Y ACTUALIZAR", True),
        ("   Guardar este archivo. Subir en el dashboard → campo 'Tabla de Capacidades'.", False),
        ("   Click 'Actualizar Reporte' para regenerar.", False),
    ]
    for ri, (txt, bold) in enumerate(lines, 1):
        c = wi.cell(ri, 1, txt)
        c.font = Font(bold=bold, size=9 if not bold else 10,
                      color="1E40AF" if bold else "1E293B", name="Segoe UI")
        c.alignment = align(h="left", indent=0 if bold else 1)
        if bold and txt:
            c.fill = fill("EFF6FF")
        wi.row_dimensions[ri].height = 15

    wb.save(OUTPUT)
    print(f"✅ Template guardado → {OUTPUT}")
    print(f"   {len(ROWS)} filas | "
          f"Total piso: {sum(r[4] for r in ROWS if r[4])} bandejas | "
          f"Total rack: {sum(r[5] for r in ROWS if r[5])} posiciones")


if __name__ == "__main__":
    write_template()
